import datetime
import locale
import os
import csv
from config.config_manager import config_manager
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Configurar el locale para formateo de moneda
try:
    locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252')
    except locale.Error:
        pass

def format_currency(amount):
    """Formatea un valor como moneda usando la configuración del sistema"""
    try:
        # Convertir a float primero
        if amount is None:
            amount = 0.0
        elif isinstance(amount, str):
            # Limpiar string de caracteres de moneda
            amount = amount.replace('$', '').replace('S/', '').replace(',', '').strip()
            if not amount or amount == '':
                amount = 0.0
            else:
                amount = float(amount)
        else:
            amount = float(amount)
        
        # Formatear con símbolo de soles
        return f"S/ {amount:,.2f}"
    except (ValueError, TypeError):
        # Fallback en caso de error
        return "S/ 0.00"

def format_date(date):
    """Formatea una fecha en formato dd/mm/yyyy"""
    if isinstance(date, str):
        try:
            date = datetime.datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return date
    return date.strftime("%d/%m/%Y")

def get_current_date():
    """Obtiene la fecha actual en formato yyyy-mm-dd"""
    return datetime.date.today().strftime("%Y-%m-%d")

def export_to_csv(data, filename, headers=None):
    """Exporta datos a un archivo CSV con codificación UTF-8-BOM"""
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'exports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    # Usar la codificación configurada en el sistema
    encoding = config_manager.get_report_encoding()
    
    with open(filepath, 'w', newline='', encoding=encoding) as f:
        writer = csv.writer(f, delimiter=';')  # Usar punto y coma como separador
        if headers:
            writer.writerow(headers)
        writer.writerows(data)
    
    return filepath

def calculate_subtotal(price, quantity):
    """Calcula el subtotal de un producto"""
    return float(price) * int(quantity)

def export_to_excel(data, filename, headers=None, sheet_name="Datos"):
    """Exporta datos a un archivo Excel con formato profesional"""
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'exports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Agregar encabezados si se proporcionan
    if headers:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    
    # Agregar datos
    start_row = 2 if headers else 1
    for row_idx, row_data in enumerate(data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filepath)
    return filepath

def generate_invoice_pdf(factura_data, filename):
    """Genera un PDF de factura con formato profesional"""
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'exports', filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Centrado
        textColor=colors.darkblue
    )
    
    # Estilo para información de la empresa
    company_style = ParagraphStyle(
        'CompanyStyle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,  # Centrado
        spaceAfter=20
    )
    
    # Título de la factura
    story.append(Paragraph("FACTURA", title_style))
    
    # Información de la empresa
    company_info = """<b>DCORELP</b><br/>
    Sistema de Gestión de Ventas<br/>
    Tel: (01) 123-4567<br/>
    Email: info@dcorelp.com"""
    story.append(Paragraph(company_info, company_style))
    story.append(Spacer(1, 20))
    
    # Información de la factura
    invoice_info = [
        ['Número de Factura:', factura_data.get('numero', 'N/A')],
        ['Fecha:', factura_data.get('fecha', 'N/A')],
        ['Estado:', factura_data.get('estado', 'N/A')],
        ['Cliente:', factura_data.get('cliente', 'N/A')],
        ['Empleado:', factura_data.get('empleado', 'N/A')]
    ]
    
    invoice_table = Table(invoice_info, colWidths=[2*inch, 3*inch])
    invoice_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(invoice_table)
    story.append(Spacer(1, 30))
    
    # Detalles financieros
    financial_data = [
        ['Concepto', 'Monto'],
        ['Subtotal:', format_currency(factura_data.get('subtotal', 0))],
        ['Impuesto:', format_currency(factura_data.get('impuesto', 0))],
        ['Total:', format_currency(factura_data.get('total', 0))]
    ]
    
    financial_table = Table(financial_data, colWidths=[3*inch, 2*inch])
    financial_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(financial_table)
    story.append(Spacer(1, 30))
    
    # Observaciones si existen
    if factura_data.get('observaciones'):
        obs_style = ParagraphStyle(
            'ObsStyle',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20
        )
        story.append(Paragraph(f"<b>Observaciones:</b> {factura_data.get('observaciones')}", obs_style))
    
    # Pie de página
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,  # Centrado
        textColor=colors.grey
    )
    story.append(Spacer(1, 50))
    story.append(Paragraph("Gracias por su preferencia", footer_style))
    story.append(Paragraph(f"Documento generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", footer_style))
    
    doc.build(story)
    return filepath